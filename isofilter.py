#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Class IsoFilter version created 3/27/17 containing functions to calculate mean, 
standard deviation, total counts, and filter data outside of specified 
range. 

Class chem_blank version created 4/19/17 contains functions to calculate mean, counts, 
and relative error of chem blank data.

Created by Julia Nissen and Nick Wang, 2017.

"""

import openpyxl
import numpy as np


class IsoFilter():
    def __init__(self, filename,columnletter,filternumber): # input filename and columnletter as strings
        self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
        self.filename = str(filename)
        self.filternumber = int(filternumber)
        self.workbook = openpyxl.load_workbook(self.filename)
        self.ws = self.workbook.active
        self.totalCounts = 0
        self.mean = 0 
        self.filteredMean = 0
        self.err = 0
        self.criteria = 0
        self.totalCounts_filt = 0
        self.standdev = 0
    
    def getMean(self):
        """
        Code works row by row through specified Excel column, and calculates total mean
        """
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value or cell.value == 0
                if value:
                    outlist.append(cell.value)
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float) 
        self.mean = np.nanmean(a = outarray)
        return self.mean

    def getStanddev(self):
        """
        Code works row by row through specified Excel column, and calculates standard deviation
        """
        outlist = []
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value or cell.value == 0
                if value:
                    outlist.append(value)
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float) 
        self.standdev = np.nanstd(a = outarray, ddof = 1)
        return self.standdev
    
    def getCounts(self):
        """
        Code works row by row through specified Excel Column, and determines total number of values present (i.e. cycles)
        """
        total_counts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value or cell.value == 0
                if value:
                    total_counts +=1

        self.totalCounts = total_counts
        return self.totalCounts
        
    def Filtered_mean(self, mean, standdev, counts):
        """
        Code works row by row through specified Excel column, deletes entries that are outside of specified range, 
        and calculates resulting mean
        """
        self.mean = mean
        self.standdev = standdev
        self.totalCounts = counts
        self.standerr = (self.standdev / (self.totalCounts**0.5))
        self.criteria = self.filternumber * self.standerr
        outlist = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value or cell.value == 0
                if value:
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                        outcounts += 1
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        self.filteredMean = np.nanmean(a = outarray)
        return self.filteredMean
    
    def Filtered_err(self, mean, standdev, counts):
        """
        Code works row by row through specified Excel column, deletes entries that are outside of specified range, 
        and calculates resulting 2s counting stantistics error
        """
        self.mean = mean
        self.standdev = standdev
        self.totalCounts = counts
        self.standerr = (self.standdev / (self.totalCounts**0.5))
        self.criteria = self.filternumber * self.standerr
        outlist = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value or cell.value == 0
                if value:
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                        outcounts += 1
                else:
                    outlist.append(np.nan)
        outarray = np.array(outlist, dtype = np.float)
        outstanddev = np.nanstd(a=outarray, ddof = 1)
        self.err = 2 * (outstanddev / (outcounts ** 0.5))
        return self.err
    
    def Filtered_counts(self, mean, standdev, counts):
        """
        Code works row by row through specified Excel column, deletes entries that are outside of specified range, 
        and determines total number of values remaining (i.e. filtered cycles)
        """
        self.mean = mean
        self.standdev = standdev
        self.totalCounts = counts
        self.standerr = (self.standdev / (self.totalCounts**0.5))
        self.criteria = self.filternumber * self.standerr
        outlist = []
        outcounts = 0
        for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
            for cell in row:
                value = cell.value or cell.value == 0
                if value:
                    if abs(value - self.mean) > self.criteria:
                        outlist.append(np.nan)
                    else:
                        outlist.append(value)
                        outcounts += 1
                else:
                    outlist.append(np.nan)
        self.totalCounts_filt = outcounts
        return self.totalCounts_filt
    
class chem_blank():
    
        def __init__(self,filename, columnletter, int_time):
            self.column = str(columnletter)+'{}:'+str(columnletter)+'{}'
            self.filename = str(filename)
            self.workbook = openpyxl.load_workbook(self.filename)
            self.ws = self.workbook.active
            
            int_time = str(int_time)
            
            int_dictionary = {"229":0.131, "230":1.049, "232":0.262, "233":0.131, "234":1.049,
                              "235": 0.262, "236":0.131, "238": 0.262}
            
            if int_time in int_dictionary:
                self.inttime = int_dictionary[int_time]
            else: print "Int_time not available"
                      
        def calc(self):
            """
            Code calculates the mean, total cycles, and 2s counting statistics error of chem blanks, in order
            to use in the Age Calculation
            """
            outlist = []
            outcounts = 0
            for row in self.ws.iter_rows(self.column.format(2, self.ws.max_row - 8)):
                for cell in row:
                    if cell.value: 
                        value = cell.value
                        outlist.append(value)
                        outcounts += 1
                    elif cell.value == 0:
                        value= 0.00
                        outlist.append(value)
                        outcounts +=1
                    else: outlist.append(np.nan)
            outarray = np.array(outlist, dtype = np.float)
            self.mean = np.nanmean(a = outarray)
            standdev = np.nanstd(a = outarray, ddof = 1)
            self.counts = outcounts
            err_abs =  2 * standdev/((self.counts)**0.5)
            err_rel_option1 = err_abs/self.mean
            err_rel_option2 = 2/((self.mean * self.counts*self.inttime)**0.5)
            self.err_rel = max(err_rel_option1, err_rel_option2)
            
            lst_Chem = [self.mean, self.counts, self.err_rel]
        
            return lst_Chem



